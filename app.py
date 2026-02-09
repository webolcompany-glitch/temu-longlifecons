import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO

st.set_page_config(page_title="Generatore Listino Temu", layout="wide")
st.title("🛒 Generatore Listino Temu")

# =========================
# FUNZIONI DI SUPPORTO
# =========================

def clean_outgoods(code):
    if pd.isna(code):
        return ""
    return str(code).split("_")[0]

def formato_label(fmt, sku=""):
    try:
        fmt = int(fmt)
    except:
        return ""
    if 1 <= fmt <= 6:
        return f"{fmt}x1L"
    if "tan" in str(sku).lower():
        return f"Tanica da {fmt}L"
    if fmt == 4:
        return "Tanica 4L"
    if fmt == 20:
        return "Tanica 20L"
    if fmt == 55:
        return "Fustino 55L"
    if fmt == 205:
        return "Fusto 205L"
    return f"{fmt}L"

def bullet_formato(fmt, sku=""):
    try:
        fmt = int(fmt)
    except:
        return ""
    if 1 <= fmt <= 6:
        return f"confezione da {fmt}x1L"
    if "tan" in str(sku).lower():
        return f"Tanica da {fmt}L"
    return f"{fmt}L"

def capacita_quantita(fmt):
    try:
        fmt = int(fmt)
    except:
        return ("", "")
    if 1 <= fmt <= 6:
        return ("1", str(fmt))
    return (str(fmt), "1")

def produttore(marca):
    if str(marca).upper().strip() == "TAMOIL":
        return "TAMOIL ITALIA S.P.A."
    return "Long life consulting s.r.l."

def nome_articolo(row):
    parts = [
        str(row.get("Sottocategoria", "")),
        str(row.get("Viscosità", "")),
        str(row.get("Marca", "")),
        str(row.get("ACEA", "")),
        formato_label(row.get("Formato (L)", ""), row.get("Sku", "")),
        str(row.get("Utilizzo", ""))
    ]
    return " ".join([p for p in parts if p])

# =========================
# UPLOAD FILE
# =========================

modello_file = st.file_uploader("📤 Carica il file modello Temu (.xlsx)", type=["xlsx"])
input_file = st.file_uploader("📤 Carica il tuo listino di input (.xlsx)", type=["xlsx"])

if modello_file and input_file:
    # Leggi il file modello Temu
    wb = load_workbook(modello_file)
    if "Template" not in wb.sheetnames:
        st.error("Il file modello non contiene il foglio 'Template'")
    else:
        ws = wb["Template"]

        # Leggi il file input
        df = pd.read_excel(input_file)
        st.subheader("Anteprima file input")
        st.dataframe(df.head())

        # Mappa header del foglio Template
        header = [cell.value for cell in ws[1]]
        col_map = {name: idx for idx, name in enumerate(header)}

        # Scrivi riga per riga
        for i, row in df.iterrows():
            r = i + 2  # perché Excel è 1-based e header è riga 1
            ws.cell(r, col_map.get("Categoria", 0)+1, 20416)
            ws.cell(r, col_map.get("Nome dell'Articolo", 0)+1, nome_articolo(row))
            ws.cell(r, col_map.get("outGoodsSn", 0)+1, clean_outgoods(row.get("Codice prodotto", "")))
            ws.cell(r, col_map.get("outSkuSn", 0)+1, row.get("Sku", ""))
            ws.cell(r, col_map.get("Aggiorna o aggiungi", 0)+1, "Aggiorna/Aggiungi nuovo")
            ws.cell(r, col_map.get("Marca", 0)+1, row.get("Marca", ""))
            ws.cell(r, col_map.get("Descrizione dell'articolo", 0)+1, row.get("Descrizione", ""))
            ws.cell(r, col_map.get("Punto elenco", 0)+1, "LONG LIFE CONSULTING: azienda italiana specializzata nel settore dei lubrificanti per autovetture, motocicli, industriali, agricoli e nautici.")
            # Punto elenco 2
            pe2_idx = [idx for idx, name in enumerate(header) if name=="Punto elenco"]
            if len(pe2_idx) > 1:
                ws.cell(r, pe2_idx[1]+1, row.get("Descrizione breve", ""))
            # Punto elenco 3
            if len(pe2_idx) > 2:
                ws.cell(r, pe2_idx[2]+1, bullet_formato(row.get("Formato (L)", ""), row.get("Sku", "")))
            # Punto elenco 4
            if len(pe2_idx) > 3:
                ws.cell(r, pe2_idx[3]+1, "SPECIFICHE TECNICHE: trovi le specifiche tecniche ben visibili sulle foto mostrate in inserzione.")

            # URL immagini dettagli (7 colonne)
            for j in range(1, 8):
                col_name = f"URL delle immagini dei dettagli"
                idxs = [idx for idx, name in enumerate(header) if name==col_name]
                if j-1 < len(idxs):
                    ws.cell(r, idxs[j-1]+1, row.get(f"Img {j}", ""))

            # URL immagini SKU (10 colonne)
            sku_cols = [idx for idx, name in enumerate(header) if name=="URL immagini SKU"]
            for k in range(min(10, len(sku_cols))):
                ws.cell(r, sku_cols[k]+1, row.get(f"Img {k+1}", ""))

            # Capacità e Quantità
            cap, qty = capacita_quantita(row.get("Formato (L)", ""))
            ws.cell(r, col_map.get("Capacità", 0)+1, cap)
            ws.cell(r, col_map.get("Quantità", 0)+1, qty)

            # Prezzi
            prezzo_base = row.get("Prezzo Marketplace", 0)
            ws.cell(r, col_map.get("Prezzo base - EUR", 0)+1, round((prezzo_base/1.22)*0.85,2))
            ws.cell(r, col_map.get("Prezzo di listino - EUR", 0)+1, prezzo_base)

            # Peso pacco e dimensioni
            ws.cell(r, col_map.get("Peso pacco - g", 0)+1, int(row.get("Formato (L)", 0)*1000))
            ws.cell(r, col_map.get("Lunghezza - cm", 0)+1, 25)
            ws.cell(r, col_map.get("Larghezza - cm", 0)+1, 25)
            ws.cell(r, col_map.get("Altezza - cm", 0)+1, 25)

            # Produttore
            ws.cell(r, col_map.get("Produttore", 0)+1, produttore(row.get("Marca", "")))
            ws.cell(r, col_map.get("Persona responsabile per l'UE.", 0)+1, "LONG LIFE CONSULTING S.R.L.")

        # Salvataggio
        buffer = BytesIO()
        wb.save(buffer)
        st.success("✅ File Temu generato correttamente")
        st.download_button(
            "⬇️ Scarica file Temu pronto",
            data=buffer.getvalue(),
            file_name="listino_temu.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
